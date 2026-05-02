import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import datetime
import ccxt
import pandas as pd
import numpy as np
import json
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ============================================================
# 1. STRATÉGIES PRÉDÉFINIES (EXEMPLES)
# ============================================================
def strategy_ema_crossover(df, fast=9, slow=21):
    df = df.copy()
    df['ema_fast'] = df['close'].ewm(span=fast, adjust=False).mean()
    df['ema_slow'] = df['close'].ewm(span=slow, adjust=False).mean()
    df['buy_signal'] = (df['ema_fast'] > df['ema_slow']) & (df['ema_fast'].shift(1) <= df['ema_slow'].shift(1))
    df['sell_signal'] = (df['ema_fast'] < df['ema_slow']) & (df['ema_fast'].shift(1) >= df['ema_slow'].shift(1))
    df['signal_text'] = ''
    df.loc[df['buy_signal'], 'signal_text'] = '🟢 ACHAT (EMA cross)'
    df.loc[df['sell_signal'], 'signal_text'] = '🔴 VENTE (EMA cross)'
    return df

def strategy_rsi(df, period=14, overbought=70, oversold=30):
    df = df.copy()
    delta = df['close'].diff()
    gain = delta.where(delta > 0, 0)
    loss = -delta.where(delta < 0, 0)
    avg_gain = gain.rolling(window=period, min_periods=period).mean()
    avg_loss = loss.rolling(window=period, min_periods=period).mean()
    rs = avg_gain / avg_loss
    df['rsi'] = 100 - (100 / (1 + rs))
    df['signal_text'] = ''
    df.loc[df['rsi'] < oversold, 'signal_text'] = '🟢 ACHAT (RSI survendu)'
    df.loc[df['rsi'] > overbought, 'signal_text'] = '🔴 VENTE (RSI suracheté)'
    return df

def strategy_bollinger_bands(df, period=20, std_dev=2):
    df = df.copy()
    df['sma'] = df['close'].rolling(window=period).mean()
    df['std'] = df['close'].rolling(window=period).std()
    df['upper'] = df['sma'] + std_dev * df['std']
    df['lower'] = df['sma'] - std_dev * df['std']
    df['signal_text'] = ''
    df.loc[df['close'] <= df['lower'], 'signal_text'] = '🟢 ACHAT (touché bande basse)'
    df.loc[df['close'] >= df['upper'], 'signal_text'] = '🔴 VENTE (touché bande haute)'
    return df

def strategy_macd(df, fast=12, slow=26, signal=9):
    df = df.copy()
    ema_fast = df['close'].ewm(span=fast, adjust=False).mean()
    ema_slow = df['close'].ewm(span=slow, adjust=False).mean()
    df['macd'] = ema_fast - ema_slow
    df['signal_line'] = df['macd'].ewm(span=signal, adjust=False).mean()
    df['buy_cross'] = (df['macd'] > df['signal_line']) & (df['macd'].shift(1) <= df['signal_line'].shift(1))
    df['sell_cross'] = (df['macd'] < df['signal_line']) & (df['macd'].shift(1) >= df['signal_line'].shift(1))
    df['signal_text'] = ''
    df.loc[df['buy_cross'], 'signal_text'] = '🟢 ACHAT (MACD croisement haussier)'
    df.loc[df['sell_cross'], 'signal_text'] = '🔴 VENTE (MACD croisement baissier)'
    return df

PREDEFINED_STRATEGIES = {
    "EMA Crossover (9/21)": strategy_ema_crossover,
    "RSI (14, 30/70)": strategy_rsi,
    "Bollinger Bands (20, 2)": strategy_bollinger_bands,
    "MACD (12,26,9)": strategy_macd,
}

# ============================================================
# 2. FONCTIONS DE RÉCUPÉRATION DE DONNÉES (ccxt)
# ============================================================
def fetch_ohlcv(exchange_name, symbol, timeframe, since, limit=1000):
    if exchange_name == 'binance':
        exchange = ccxt.binance()
    elif exchange_name == 'coinbase':
        exchange = ccxt.coinbase()
    else:
        raise ValueError("Exchange non supporté")

    if not exchange.has['fetchOHLCV']:
        raise Exception(f"L'exchange {exchange_name} ne supporte pas fetchOHLCV")

    all_ohlcv = []
    current_since = since
    while True:
        try:
            ohlcv = exchange.fetch_ohlcv(symbol, timeframe, since=current_since, limit=limit)
            if not ohlcv:
                break
            all_ohlcv.extend(ohlcv)
            current_since = ohlcv[-1][0] + 1
            if len(ohlcv) < limit:
                break
        except Exception as e:
            raise Exception(f"Erreur lors du téléchargement : {e}")

    if not all_ohlcv:
        return pd.DataFrame()
    df = pd.DataFrame(all_ohlcv, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume'])
    df['datetime'] = pd.to_datetime(df['timestamp'], unit='ms')
    df.set_index('datetime', inplace=True)
    return df

# ============================================================
# 3. APPLICATION TKINTER
# ============================================================
class CandlestickDownloaderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Téléchargeur de données Candlestick + Stratégies Python")
        self.root.geometry("1000x900")

        # Variables
        self.exchange_var = tk.StringVar(value="binance")
        self.symbol_var = tk.StringVar(value="BTC/USDT")
        self.interval_var = tk.StringVar(value="1h")
        self.date_type_var = tk.StringVar(value="range")
        self.start_datetime_var = tk.StringVar()
        self.end_datetime_var = tk.StringVar()
        self.destination_folder = tk.StringVar()
        self.strategy_var = tk.StringVar(value=list(PREDEFINED_STRATEGIES.keys())[0])
        self.custom_python_code = tk.StringVar()

        self.intervals = ['1m', '3m', '5m', '15m', '30m', '1h', '2h', '4h', '6h', '8h', '12h', '1d', '3d', '1w', '1M']

        self.create_widgets()

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Section 1
        group1 = ttk.LabelFrame(main_frame, text="1. Source des données", padding="5")
        group1.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(group1, text="Binance", variable=self.exchange_var, value="binance").pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(group1, text="Coinbase", variable=self.exchange_var, value="coinbase").pack(side=tk.LEFT, padx=10)
        ttk.Label(group1, text="Symbole (ex: BTC/USDT) :").pack(side=tk.LEFT, padx=10)
        ttk.Entry(group1, textvariable=self.symbol_var, width=15).pack(side=tk.LEFT)

        # Section 2
        group2 = ttk.LabelFrame(main_frame, text="2. Période", padding="5")
        group2.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(group2, text="Plage glissante (jours)", variable=self.date_type_var, value="range", command=self.toggle_date_type).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(group2, text="Dates précises (UTC)", variable=self.date_type_var, value="dates", command=self.toggle_date_type).pack(side=tk.LEFT, padx=10)

        self.range_frame = ttk.Frame(group2)
        ttk.Label(self.range_frame, text="Nombre de jours :").pack(side=tk.LEFT, padx=5)
        self.days_spinbox = tk.Spinbox(self.range_frame, from_=1, to=3650, width=8)
        self.days_spinbox.delete(0, tk.END)
        self.days_spinbox.insert(0, "30")
        self.days_spinbox.pack(side=tk.LEFT)

        self.dates_frame = ttk.Frame(group2)
        ttk.Label(self.dates_frame, text="Date début (YYYY-MM-DD HH:MM:SS) :").pack(side=tk.LEFT, padx=5)
        ttk.Entry(self.dates_frame, textvariable=self.start_datetime_var, width=20).pack(side=tk.LEFT)
        ttk.Label(self.dates_frame, text="Date fin :").pack(side=tk.LEFT, padx=5)
        ttk.Entry(self.dates_frame, textvariable=self.end_datetime_var, width=20).pack(side=tk.LEFT)

        self.range_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.dates_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.dates_frame.pack_forget()

        # Section 3
        group3 = ttk.LabelFrame(main_frame, text="3. Intervalle des bougies", padding="5")
        group3.pack(fill=tk.X, pady=5)
        ttk.Label(group3, text="Choisissez l'intervalle :").pack(side=tk.LEFT, padx=5)
        ttk.Combobox(group3, textvariable=self.interval_var, values=self.intervals, state="readonly", width=10).pack(side=tk.LEFT)

        # Section 4
        group4 = ttk.LabelFrame(main_frame, text="4. Dossier de destination", padding="5")
        group4.pack(fill=tk.X, pady=5)
        ttk.Entry(group4, textvariable=self.destination_folder, width=50).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(group4, text="Parcourir", command=self.browse_folder).pack(side=tk.LEFT, padx=5)

        # Section 5 : Choix de la stratégie
        group5 = ttk.LabelFrame(main_frame, text="5. Stratégie à appliquer", padding="5")
        group5.pack(fill=tk.BOTH, expand=True, pady=5)

        ttk.Label(group5, text="Stratégie prédéfinie :").grid(row=0, column=0, padx=5, sticky=tk.W)
        ttk.Combobox(group5, textvariable=self.strategy_var, values=list(PREDEFINED_STRATEGIES.keys()), state="readonly", width=40).grid(row=0, column=1, padx=5, sticky=tk.W)

        ttk.Label(group5, text="OU code Python personnalisé (fonction custom_strategy(df)) :").grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky=tk.W)
        self.python_editor = tk.Text(group5, height=15, wrap=tk.WORD, font=("Courier", 9))
        self.python_editor.grid(row=2, column=0, columnspan=3, padx=5, pady=5, sticky=tk.W+tk.E)
        self.python_editor.insert(tk.END, "# Exemple :\n# def custom_strategy(df):\n#     df = df.copy()\n#     df['signal_text'] = ''\n#     df.loc[df['close'] > df['open'], 'signal_text'] = '🟢 ACHAT'\n#     return df")

        # Bouton pour appliquer la stratégie personnalisée
        self.apply_custom_btn = ttk.Button(group5, text="Utiliser le code personnalisé ci-dessus", command=self.use_custom_strategy)
        self.apply_custom_btn.grid(row=3, column=0, padx=5, pady=5)

        # Bouton principal de téléchargement
        self.download_btn = ttk.Button(main_frame, text="Télécharger les données et appliquer la stratégie", command=self.start_download, width=50)
        self.download_btn.pack(pady=10)

        # Barre de progression
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.pack(fill=tk.X, pady=5)

        # Zone de statut
        self.status_text = tk.Text(main_frame, height=10, state=tk.DISABLED, wrap=tk.WORD)
        self.status_text.pack(fill=tk.BOTH, expand=True, pady=5)
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.status_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.status_text.configure(yscrollcommand=scrollbar.set)

    def toggle_date_type(self):
        if self.date_type_var.get() == "range":
            self.range_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
            self.dates_frame.pack_forget()
        else:
            self.dates_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
            self.range_frame.pack_forget()

    def browse_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.destination_folder.set(folder)

    def use_custom_strategy(self):
        code = self.python_editor.get(1.0, tk.END).strip()
        if not code:
            messagebox.showerror("Erreur", "Le code personnalisé est vide.")
            return
        try:
            # Tester la compilation
            exec_globals = {'pd': pd, 'np': np}
            exec(code, exec_globals)
            if 'custom_strategy' not in exec_globals:
                raise Exception("Le code doit contenir une fonction 'custom_strategy(df)'.")
            self.custom_strategy_func = exec_globals['custom_strategy']
            self.log("✅ Stratégie personnalisée chargée avec succès.")
        except Exception as e:
            self.log(f"❌ Erreur dans le code personnalisé : {e}")
            messagebox.showerror("Erreur", f"Erreur dans le code : {e}")

    def log(self, message):
        self.status_text.config(state=tk.NORMAL)
        self.status_text.insert(tk.END, message + "\n")
        self.status_text.see(tk.END)
        self.status_text.config(state=tk.DISABLED)
        self.root.update_idletasks()

    def start_download(self):
        if not self.destination_folder.get():
            messagebox.showerror("Erreur", "Veuillez sélectionner un dossier de destination.")
            return
        self.download_btn.config(state=tk.DISABLED)
        self.progress.start(10)
        thread = threading.Thread(target=self.download_and_process)
        thread.daemon = True
        thread.start()

    def download_and_process(self):
        try:
            exchange = self.exchange_var.get()
            symbol = self.symbol_var.get()
            timeframe = self.interval_var.get()

            # Période
            if self.date_type_var.get() == "range":
                days = int(self.days_spinbox.get())
                end_dt = datetime.datetime.now(datetime.timezone.utc)
                start_dt = end_dt - datetime.timedelta(days=days)
                since = int(start_dt.timestamp() * 1000)
                self.log(f"Période : du {start_dt} au {end_dt} (UTC)")
            else:
                start_str = self.start_datetime_var.get()
                end_str = self.end_datetime_var.get()
                start_dt = datetime.datetime.strptime(start_str, "%Y-%m-%d %H:%M:%S").replace(tzinfo=datetime.timezone.utc)
                end_dt = datetime.datetime.strptime(end_str, "%Y-%m-%d %H:%M:%S").replace(tzinfo=datetime.timezone.utc)
                since = int(start_dt.timestamp() * 1000)
                self.log(f"Période : du {start_dt} au {end_dt} (UTC)")

            self.log(f"Téléchargement des données {symbol} sur {exchange} (intervalle {timeframe})...")
            df = fetch_ohlcv(exchange, symbol, timeframe, since)
            if df.empty:
                self.log("Aucune donnée trouvée pour cette période.")
                return
            self.log(f"Données téléchargées : {len(df)} bougies")

            # Appliquer la stratégie
            strategy_name = self.strategy_var.get()
            if hasattr(self, 'custom_strategy_func') and self.custom_strategy_func:
                self.log("Application de la stratégie personnalisée...")
                df_result = self.custom_strategy_func(df)
            elif strategy_name in PREDEFINED_STRATEGIES:
                self.log(f"Application de la stratégie prédéfinie : {strategy_name}")
                df_result = PREDEFINED_STRATEGIES[strategy_name](df)
            else:
                raise Exception("Aucune stratégie valide sélectionnée.")

            if self.date_type_var.get() == "dates":
                df_result = df_result[df_result.index <= end_dt]

            if 'signal_text' not in df_result.columns:
                df_result['signal_text'] = ''
            export_df = df_result[['open', 'high', 'low', 'close', 'volume', 'signal_text']].copy()
            export_df.reset_index(inplace=True)
            export_df.rename(columns={'datetime': 'Date/Heure (UTC)'}, inplace=True)

            base_name = f"{symbol.replace('/', '_')}_{timeframe}_{exchange}_{start_dt.strftime('%Y%m%d_%H%M%S')}_to_{end_dt.strftime('%Y%m%d_%H%M%S')}"
            file_path = f"{self.destination_folder.get()}/{base_name}.xlsx"

            self.save_to_excel(export_df, file_path, strategy_name if not hasattr(self, 'custom_strategy_func') else "Personnalisée", start_dt, end_dt)
            self.log(f"✅ Fichier sauvegardé : {file_path}")

        except Exception as e:
            self.log(f"❌ Erreur : {str(e)}")
            messagebox.showerror("Erreur", str(e))
        finally:
            self.progress.stop()
            self.download_btn.config(state=tk.NORMAL)

    def save_to_excel(self, df, filepath, strategy_name, start_dt, end_dt):
        wb = Workbook()
        ws = wb.active
        ws.title = "Candles + Signaux"

        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        signal_col_idx = headers.index('signal_text') + 1 if 'signal_text' in headers else None
        if signal_col_idx:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=signal_col_idx)
                val = cell.value
                if val and "ACHAT" in str(val):
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                elif val and "VENTE" in str(val):
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

        ws_params = wb.create_sheet("Paramètres")
        params = [
            ("Exchange", self.exchange_var.get()),
            ("Symbole", self.symbol_var.get()),
            ("Intervalle", self.interval_var.get()),
            ("Date début", start_dt.strftime("%Y-%m-%d %H:%M:%S")),
            ("Date fin", end_dt.strftime("%Y-%m-%d %H:%M:%S")),
            ("Stratégie", strategy_name),
            ("Nombre de bougies", len(df))
        ]
        for i, (key, val) in enumerate(params, start=1):
            ws_params.cell(row=i, column=1, value=key).font = Font(bold=True)
            ws_params.cell(row=i, column=2, value=val)

        wb.save(filepath)

# ============================================================
# 4. LANCEMENT
# ============================================================
if __name__ == "__main__":
    root = tk.Tk()
    app = CandlestickDownloaderApp(root)
    root.mainloop()